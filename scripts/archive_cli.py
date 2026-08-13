# -*- coding: utf-8 -*-
"""
档案电子目录 · 一体化解析 CLI（通用模板）

把「解析源文件 → 规范化 → 构建成品 → 并入总目录 → 校验」集成到一个命令。
全宗号等参数化，复制到目标项目后改 QUANZONG 即可。

依赖：openpyxl, norm_common

用法：
    # 新建成品（从原始源构建归档目录/卷内目录）
    python archive_cli.py build <成品.xlsx> <源文件1.xlsx> [源文件2.xlsx ...]

    # 并入新批次到已确认总目录（基于已确认成品，勿重跑旧批）
    python archive_cli.py merge <总目录.xlsx> <新文件.xlsx>

    # 校验成品
    python archive_cli.py check <成品.xlsx>

    # 列出/清除黄色高亮
    python archive_cli.py highlight <成品.xlsx>            # 列出
    python archive_cli.py highlight <成品.xlsx> --clear    # 清除
"""
import re
import sys
import shutil
import argparse
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from norm_common import clean, norm_dh, norm_wh, norm_tm, norm_zrz, norm_pg, date_ok

# ---------- 常量（可按单位调整） ----------
QUANZONG = "1031"                    # 全宗号：总工会 1031 / 县委办 1001 / 统战部 1012
FONT = "宋体"
THIN = Side(style="thin")
BORDER = Border(left=THIN, top=THIN, right=THIN, bottom=THIN)
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
COL_WIDTHS = {1: 5.255, 2: 9.755, 3: 11.755, 4: 15.255, 5: 22.873,
              6: 9.627, 7: 5.127, 8: 5.127, 9: 5.127}
HEADERS = ["序号", "档 号", "文号", "责任者", "题 名", "日 期", "密级", "页数", "备注"]
FIELD_COL = {"序号": 1, "文号": 3, "责任者": 4, "题名": 5, "日期": 6, "密级": 7, "页数": 8, "备注": 9}
PERIOD_ORDER = ["永久", "30年", "10年"]


# ============================================================
# 解析
# ============================================================
def parse_archive(path):
    """解析归档文件目录（含档号），返回 [(year, period, data), ...]。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    period = None
    fieldmap = None
    items = []
    unparsed = []
    for row in ws.iter_rows(values_only=True):
        a = clean(row[0])
        if "归档文件目录索引" in a:
            continue
        m = re.search(r"保管期限[：:]?\s*(\S+)", a)
        if m:
            period = m.group(1)
            fieldmap = None
            continue
        if "档号" in a or any(c and "档号" in str(c) for c in row):
            fieldmap = {}
            for i, cell in enumerate(row):
                t = clean(cell)
                if t in ("序号", "档号", "文号", "责任者", "题名", "日期", "密级", "页数", "备注"):
                    fieldmap[t] = i
            continue
        if period is None or fieldmap is None:
            continue
        if not a.isdigit():
            continue
        data = {k: clean(row[i]) for k, i in fieldmap.items()}
        core = data.get("档号", "") + data.get("文号", "") + data.get("责任者", "") + data.get("题名", "")
        if not re.search(r"[一-鿿0-9]", core):
            continue
        dh = norm_dh(data.get("档号", ""))
        m2 = re.fullmatch(rf"{QUANZONG}-[a-z]+·(\d{{4}})-([a-z0-9]+)-(\d{{4}})", dh)
        if not m2:
            unparsed.append((period, data.get("档号", ""), data.get("题名", "")[:30]))
            continue
        year, code, no = m2.groups()
        data["档号"] = dh
        data["文号"] = norm_wh(data.get("文号", ""))
        data["题名"] = norm_tm(data.get("题名", ""))
        data["责任者"] = norm_zrz(data.get("责任者", ""))
        data["日期"] = data.get("日期", "").strip()
        data["页数"] = norm_pg(data.get("页数", ""))
        data["密级"] = data.get("密级", "").strip()
        data["备注"] = data.get("备注", "").strip()
        # 日期异常分级：9位占位→留空；非法/缺位→标记待确认
        ok, _ = date_ok(data["日期"])
        if not ok:
            if re.fullmatch(r"\d{9}", data["日期"]) and data["日期"][4:6] == "00" and data["日期"][6:8] == "00":
                data["日期"] = ""            # 9位占位（201000000）→ 留空
            else:
                data["_hl"] = set(data.get("_hl", ())) | {"日期"}   # 非法/缺位 → 高亮待确认
        items.append((year, period, data))
    return items, unparsed


# ============================================================
# 输出构建
# ============================================================
def row_height_calc(vals):
    import math
    max_lines = 1
    for col, v in enumerate(vals, 1):
        if v:
            units = sum(2 if ord(ch) > 0x2E80 else 1 for ch in str(v))
            per_line = max(1, int(COL_WIDTHS[col] * 0.8))
            max_lines = max(max_lines, math.ceil(units / per_line))
    return max(45.0, max_lines * 17 + 4)


def build_sheet(wb, sheet_name, period, rows):
    """在 wb 中新建 年度×期限 Sheet。rows=[(seq, data), ...]"""
    ws = wb.create_sheet(sheet_name)
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.merge_cells("A1:I1")
    ws["A1"] = "档案文件目录"
    ws["A1"].font = Font(name=FONT, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 23.2
    ws.merge_cells("F2:I2")
    ws["F2"] = f"保管期限：{period}"
    ws["F2"].font = Font(name=FONT, size=12)
    ws["F2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 20.2
    ws.row_dimensions[3].height = 15.0
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = Font(name=FONT, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for r, (seq, data) in enumerate(rows, 4):
        vals = [seq, data.get("档号", ""), data.get("文号", ""), data.get("责任者", ""),
                data.get("题名", ""), data.get("日期", ""),
                data.get("密级", ""), data.get("页数", ""), data.get("备注", "")]
        ws.row_dimensions[r].height = row_height_calc(vals)
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name=FONT, size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
        for hc in data.get("_hl", set()):
            col = FIELD_COL.get(hc)
            if col:
                ws.cell(row=r, column=col).fill = YELLOW_FILL
    ws.freeze_panes = "A4"


def cmd_build(args):
    """新建成品：解析源文件 → 按 年度×期限 分 Sheet 构建。"""
    items = []
    for src in args.srcs:
        its, un = parse_archive(src)
        print(f"{src}: {len(its)} 件，档号无法解析 {len(un)} 条")
        for u in un:
            print("   !! 未解析:", u)
        items.extend(its)
    groups = {}
    for year, period, data in items:
        groups.setdefault((year, period), []).append(data)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for (year, period) in sorted(groups, key=lambda k: (int(k[0]), PERIOD_ORDER.index(k[1]) if k[1] in PERIOD_ORDER else 99)):
        rows = [(f"{i:03d}", d) for i, d in enumerate(groups[(year, period)], 1)]
        sheet_name = f"{year}{period}" if period == "永久" else f"{year}-{period}"
        build_sheet(wb, sheet_name, period, rows)
        print(f"  {sheet_name}: {len(rows)} 件")
    wb.save(args.dst)
    print(f"已保存: {args.dst}")


def cmd_merge(args):
    """并入新批次到已确认总目录（保留原 Sheet 与已确认内容，仅覆盖同名 Sheet）。"""
    backup = args.master.replace(".xlsx", "_bak.xlsx")
    shutil.copy2(args.master, backup)
    print(f"已备份: {backup}")
    items, un = parse_archive(args.src)
    print(f"{args.src}: {len(items)} 件，档号无法解析 {len(un)} 条")
    for u in un:
        print("   !! 未解析:", u)
    groups = {}
    for year, period, data in items:
        groups.setdefault((year, period), []).append(data)
    wb = openpyxl.load_workbook(args.master)
    for (year, period) in sorted(groups, key=lambda k: (int(k[0]), PERIOD_ORDER.index(k[1]) if k[1] in PERIOD_ORDER else 99)):
        rows = [(f"{i:03d}", d) for i, d in enumerate(groups[(year, period)], 1)]
        sheet_name = f"{year}{period}" if period == "永久" else f"{year}-{period}"
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        build_sheet(wb, sheet_name, period, rows)
        print(f"  {sheet_name}: {len(rows)} 件")
    wb.save(args.master)
    print(f"已并入: {args.master}")


# ============================================================
# 校验 / 高亮
# ============================================================
def cmd_check(args):
    """最终校验清单：Sheet/件数/序号/档号/高亮/日期/页数。"""
    wb = openpyxl.load_workbook(args.dst)
    problems = []
    total = 0
    all_dh = {}
    for ws in wb.worksheets:
        nums = []
        for r in ws.iter_rows(min_row=4):
            if r[0].value is None:
                continue
            nums.append(int(r[0].value))
            total += 1
            dh = r[1].value
            if dh:
                if dh in all_dh:
                    problems.append(f"{ws.title}: 档号重复 {dh}")
                all_dh[dh] = ws.title
            d = r[5].value
            if d not in (None, ""):
                ok, why = date_ok(str(d))
                if not ok:
                    problems.append(f"{ws.title}!{r[0].value}: {why}")
            p = r[7].value
            if p not in (None, "") and not re.fullmatch(r"\d{3}(-\d{3})?", str(p).strip()):
                problems.append(f"{ws.title}!{r[0].value}: 页数格式 {p!r}")
            if r[0].value is not None and r[0].value != "" and r[0].value is not None:
                pass
            for cell in r:
                if cell.fill and cell.fill.patternType:
                    problems.append(f"{ws.title}!{cell.coordinate}: 残留高亮")
        if nums:
            missing = [i for i in range(1, max(nums) + 1) if i not in nums]
            if missing:
                problems.append(f"{ws.title}: 序号缺 {missing}")
    print(f"共 {total} 件，{len(wb.sheetnames)} 个 Sheet")
    print(f"档号 {len(all_dh)} 个")
    if problems:
        print(f"\n发现 {len(problems)} 处问题:")
        for p in problems:
            print("  -", p)
    else:
        print("\n✅ 校验通过，无问题")


def cmd_highlight(args):
    """列出或清除黄色高亮。"""
    wb = openpyxl.load_workbook(args.dst)
    hits = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.fill and cell.fill.patternType:
                    hits.append((ws.title, cell.coordinate, str(cell.value)[:20] if cell.value else ""))
    if args.clear:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.fill and cell.fill.patternType:
                        cell.fill = PatternFill(fill_type=None)
        wb.save(args.dst)
        print(f"已清除 {len(hits)} 处高亮: {args.dst}")
    else:
        print(f"黄色高亮 {len(hits)} 处:")
        for h in hits:
            print("  ", h)


# ============================================================
# main
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="档案电子目录一体化 CLI")
    sub = parser.add_subparsers(dest="cmd", required=True)
    p_build = sub.add_parser("build", help="新建成品")
    p_build.add_argument("dst")
    p_build.add_argument("srcs", nargs="+")
    p_build.set_defaults(fn=cmd_build)
    p_merge = sub.add_parser("merge", help="并入新批次")
    p_merge.add_argument("master")
    p_merge.add_argument("src")
    p_merge.set_defaults(fn=cmd_merge)
    p_check = sub.add_parser("check", help="校验成品")
    p_check.add_argument("dst")
    p_check.set_defaults(fn=cmd_check)
    p_hl = sub.add_parser("highlight", help="列出/清除高亮")
    p_hl.add_argument("dst")
    p_hl.add_argument("--clear", action="store_true")
    p_hl.set_defaults(fn=cmd_highlight)
    args = parser.parse_args()
    args.fn(args)


if __name__ == "__main__":
    main()
